#!/usr/bin/env python
# -*- coding: utf-8 -*-
#
#   get-feishu-table.py --- Description
#
#   Copyright (C) 2023, Schspa, all rights reserved.
#
#   This program is free software; you can redistribute it and/or modify
#   it under the terms of the GNU General Public License as published by
#   the Free Software Foundation; either version 2 of the License, or
#   (at your option) any later version.
#
#   This program is distributed in the hope that it will be useful,
#   but WITHOUT ANY WARRANTY; without even the implied warranty of
#   MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
#   GNU General Public License for more details.
#
#   You should have received a copy of the GNU General Public License
#   along with this program.  If not, see <http://www.gnu.org/licenses/>.
#

import os
import sys
import time
import hashlib
import requests
import keyring
import json
from loguru import logger
import http.server
import socketserver
from urllib.parse import urlparse, parse_qs
from auth import FeishuAuth
import threading
import click

def column_number_to_name(column_number):
    if column_number <= 0:
        raise ValueError("Column number should be greater than 0")

    column_name = ""
    while column_number > 0:
        remainder = (column_number - 1) % 26  # 0 to 25 for A to Z
        column_name = chr(65 + remainder) + column_name
        column_number = (column_number - 1) // 26  # Integer division for next digit

    return column_name

def get_cell_name(row, col):
    return '{:s}{:d}'.format(column_number_to_name(col), row)

class SpreadSheets(FeishuAuth):
    """
    电子表格概述: https://open.feishu.cn/document/ukTMukTMukTM/uATMzUjLwEzM14CMxMTN/overview
    """

    def __init__(self, feishu_host, app_id, app_secret, url):
        super().__init__(feishu_host, app_id, app_secret)
        self.host = feishu_host
        self.url = url
        parsed_url = urlparse(url)
        url_path = parsed_url.path
        query_dict = parse_qs(parsed_url.query)
        # Split the URL path into parts
        path_parts = url_path.split('/')
        # Remove empty parts
        path_parts = [part for part in path_parts if part]
        self.sheet_info = {}

        if path_parts[0] == 'wiki':
            # https://open.feishu.cn/wiki/xxxxxxxxxxxx?sheet=xxx
            logger.info("Get a wiki file, try to get it's node info")
            node_info = super().request('GET', 'https://open.feishu.cn/open-apis/wiki/v2/spaces/get_node', params = {
                'obj_type': 'wiki',
                'token': path_parts[1]
            })['data']
            obj_token = node_info['node']['obj_token']
            obj_type = node_info['node']['obj_type']
            if obj_type != 'sheet':
                raise Exception
            title = node_info['node']['title']
            logger.debug(f'obj_type: { obj_type }, obj_token: { obj_token } title: { title }')
            self.sheet_token = obj_token
        else:
            self.sheet_token = path_parts[1]

        sheet_id = query_dict.get('sheet', None)
        if sheet_id is not None:
            self.sheet_id = sheet_id[0]
        else:
            self.sheet_id = None

        logger.debug(f'Create sheet object token: {self.sheet_token}, sheet_id: {self.sheet_id}')
        self.check()

    def check(self):
        '''
        get sheet basic information
        '''
        """获取电子表格信息

        https://open.feishu.cn/document/ukTMukTMukTM/uUDN04SN0QjL1QDN/sheets-v3/spreadsheet/get

        :param token: 表格的token，示例值："shtxxxxxxxxxxxxxxx"
        :return: 返回数据样例如下
                {
                    "code": 0,
                    "msg": "success",
                    "data": {
                        "spreadsheet": {
                            "title": "title",
                            "owner_id": "ou_xxxxxxxxxxxx",
                            "token": "shtxxxxxxxxxxxxxx",
                            "url": "https://bytedance.feishu.cn/sheets/shtcnmBA*****yGehy8"
                        }
                    }
                }
        """
        url = f"{self.host}/open-apis/sheets/v3/spreadsheets/{self.sheet_token}"
        return super().request('GET', url)

    @property
    def sheets(self):
        sheets = []
        url = 'https://open.feishu.cn/open-apis/sheets/v3/spreadsheets/{:s}/sheets/query'.format(self.sheet_token)
        sheet_tables = super().request("GET", url)
        return sheet_tables['data']['sheets']

    def get_sheet_info(self, sheet_id = None):
        if sheet_id is None:
            sheet_id = self.sheet_id
        sheet_info = self.sheet_info.get(sheet_id, None)
        if sheet_info is None:
            ranges = '{:s}!A:AM'.format(self.sheet_id)
            url = 'https://open.feishu.cn/open-apis/sheets/v3/spreadsheets/{:s}/sheets/{:s}'.format(self.sheet_token, sheet_id)
            sheet_info = super().request("GET", url)
            self.sheet_info[sheet_id] = sheet_info

        return sheet_info

    def get_max_row(self, sheet_id = None):
        '''
        {'code': 0, 'data': {'sheet': {'grid_properties': {'column_count': 10, 'frozen_column_count': 0, 'frozen_row_count': 1, 'row_count': 741}, 'hidden': False, 'index': 1, 'merges':
        [
        {'end_column_index': 0, 'end_row_index': 316, 'start_column_index': 0, 'start_row_index': 291},
        {'end_column_index': 0, 'end_row_index': 393, 'start_column_index': 0, 'start_row_index': 375},
        {'end_column_index': 0, 'end_row_index': 374, 'start_column_index': 0, 'start_row_index': 344}, {'end_column_index': 0, 'end_row_index': 343, 'start_column_index': 0, 'start_row_index': 317}, {'end_column_index': 0, 'end_row_index': 290, 'start_column_index': 0, 'start_row_index': 269}, {'end_column_index': 0, 'end_row_index': 268, 'start_column_index': 0, 'start_row_index': 251}, {'end_column_index': 0, 'end_row_index': 250, 'start_column_index': 0, 'start_row_index': 207}, {'end_column_index': 0, 'end_row_index': 153, 'start_column_index': 0, 'start_row_index': 136}, {'end_column_index': 0, 'end_row_index': 135, 'start_column_index': 0, 'start_row_index': 130}, {'end_column_index': 0, 'end_row_index': 129, 'start_column_index': 0, 'start_row_index': 118}, {'end_column_index': 0, 'end_row_index': 117, 'start_column_index': 0, 'start_row_index': 106}, {'end_column_index': 0, 'end_row_index': 105, 'start_column_index': 0, 'start_row_index': 95}, {'end_column_index': 0, 'end_row_index': 94, 'start_column_index': 0, 'start_row_index': 85}, {'end_column_index': 0, 'end_row_index': 84, 'start_column_index': 0, 'start_row_index': 64}, {'end_column_index': 0, 'end_row_index': 63, 'start_column_index': 0, 'start_row_index': 43}, {'end_column_index': 0, 'end_row_index': 42, 'start_column_index': 0, 'start_row_index': 22}, {'end_column_index': 0, 'end_row_index': 21, 'start_column_index': 0, 'start_row_index': 1}, {'end_column_index': 0, 'end_row_index': 182, 'start_column_index': 0, 'start_row_index': 154}, {'end_column_index': 0, 'end_row_index': 206, 'start_column_index': 0, 'start_row_index': 183}, {'end_column_index': 0, 'end_row_index': 458, 'start_column_index': 0, 'start_row_index': 454}, {'end_column_index': 0, 'end_row_index': 473, 'start_column_index': 0, 'start_row_index': 459}, {'end_column_index': 0, 'end_row_index': 453, 'start_column_index': 0, 'start_row_index': 449}, {'end_column_index': 0, 'end_row_index': 448, 'start_column_index': 0, 'start_row_index': 394}, {'end_column_index': 0, 'end_row_index': 596, 'start_column_index': 0, 'start_row_index': 524}, {'end_column_index': 0, 'end_row_index': 478, 'start_column_index': 0, 'start_row_index': 474}, {'end_column_index': 0, 'end_row_index': 522, 'start_column_index': 0, 'start_row_index': 479}, {'end_column_index': 0, 'end_row_index': 613, 'start_column_index': 0, 'start_row_index': 597}, {'end_column_index': 0, 'end_row_index': 632, 'start_column_index': 0, 'start_row_index': 615}, {'end_column_index': 0, 'end_row_index': 643, 'start_column_index': 0, 'start_row_index': 640}, {'end_column_index': 0, 'end_row_index': 639, 'start_column_index': 0, 'start_row_index': 633}, {'end_column_index': 0, 'end_row_index': 661, 'start_column_index': 0, 'start_row_index': 656}, {'end_column_index': 0, 'end_row_index': 672, 'start_column_index': 0, 'start_row_index': 662}, {'end_column_index': 0, 'end_row_index': 692, 'start_column_index': 0, 'start_row_index': 673}, {'end_column_index': 0, 'end_row_index': 696, 'start_column_index': 0, 'start_row_index': 693}, {'end_column_index': 0, 'end_row_index': 701, 'start_column_index': 0, 'start_row_index': 698}, {'end_column_index': 0, 'end_row_index': 712, 'start_column_index': 0, 'start_row_index': 702},
        {'end_column_index': 0, 'end_row_index': 734, 'start_column_index': 0, 'start_row_index': 713},
        {'end_column_index': 0, 'end_row_index': 655, 'start_column_index': 0, 'start_row_index': 644}
        ], 'resource_type': 'sheet', 'sheet_id': 'I1Z4dW', 'title': 'aaa'}}, 'msg': ''}
        '''
        sheet_info = self.get_sheet_info()
        grid_properties = sheet_info['data']['sheet']['grid_properties']
        return grid_properties['row_count'], grid_properties['column_count']

    def get_merge_info(self, sheet_id = None):
        sheet_info = self.get_sheet_info()
        if 'merges' in sheet_info['data']['sheet'].keys():
            return sheet_info['data']['sheet']['merges']
        else:
            return []

    def get_value(self, row, column):
        ranges = '{sheet_id:s}!{column_name:s}{row:d}:{column_name:s}{row:d}'.format(
            sheet_id = self.sheet_id,
            column_name = column_number_to_name(column),
            row = row
        )
        url = 'https://open.feishu.cn/open-apis/sheets/v2/spreadsheets/{:s}/values/{:s}'.format(self.sheet_token, ranges)
        sheet_tables = super().request("GET", url)
        return sheet_tables['data']


    def get_table_data(self):
        max_row, max_column = self.get_max_row()
        column_idx = column_number_to_name(max_column)
        logger.debug(f'max_row: {max_row} max_column: {max_column}, column_idx: {column_idx}')
        ranges = '{:s}!A:{:s}'.format(self.sheet_id, column_idx)
        url = 'https://open.feishu.cn/open-apis/sheets/v2/spreadsheets/{:s}/values/{:s}'.format(self.sheet_token, ranges)
        sheet_tables = super().request("GET", url)
        return sheet_tables['data']

    def to_excel(self, file_path):
        from openpyxl import Workbook
        wb = Workbook()
        for sheet in self.sheets:
            logger.debug(f"Process sheet: {sheet['title']} id: { sheet['sheet_id'] }")
            self.sheet_id = sheet['sheet_id']
            ws = wb.create_sheet(sheet['title'], -1)

            table_data = self.get_table_data()

            row = 0
            col = 0
            for rv in table_data['valueRange']['values']:
                row = row + 1
                col = 0
                for v in rv:
                    col = col + 1
                    cell = ws.cell(row = row, column = col)
                    try:
                        cell.value = v
                    except ValueError:
                        if isinstance(v, dict):
                            cell.value = v['text']
                        if isinstance(v, list):
                            text = ''
                            for line in v:
                                text = text + line['text']

                            cell.value = text

            # do merge cell
            for merge_info in self.get_merge_info():
                # {'end_column_index': 0, 'end_row_index': 316, 'start_column_index': 0, 'start_row_index': 291},
                start = get_cell_name(row = merge_info['start_row_index'] + 1, col = merge_info['start_column_index'] + 1)
                end = get_cell_name(row = merge_info['end_row_index'] + 1, col = merge_info['end_column_index'] + 1)
                merge_range = '{:s}:{:s}'.format(start, end)
                logger.debug(f"Merge {merge_range}")
                ws.merge_cells(merge_range)

        wb.save(file_path)

FEISHU_HOST = 'https://open.feishu.cn'

@click.command()
@click.option('-u', '--sheet-url', 'sheet_url', type = str, default = None)
@click.option('-o', '--output', 'output_file', type = click.File('wb'), default = None)
@click.option('--app_cfg', 'app_cfg', type = str, default = 'default')
@click.option('--setup', 'setup', is_flag = True, default = False)
@click.option('-v', '--verbose', is_flag=True, default = False)
def get_feishu_table(sheet_url, output_file, app_cfg, setup, verbose):
    handler = {"sink": sys.stderr, "level": "INFO"}
    if verbose:
        handler = {"sink": sys.stderr, "level": "DEBUG"}
    logger.configure(handlers=[handler])

    feishu_cfg = keyring.get_password('FeishuApp', app_cfg)
    app_id = None
    app_secret = None
    feishu_host = 'https://open.feishu.cn'
    if feishu_cfg is None or setup:
        import getpass

        app_id = getpass.getpass(f'Please input the feishu app id:')
        app_secret = getpass.getpass(f'Please input the feishu app secret:')
        feishu_host = getpass.getpass(f'Please input the feishu host url, default: {feishu_host}:')
        if feishu_host == '':
            feishu_host = FEISHU_HOST
        feishu_cfg = {
            'app_id' : app_id,
            'app_secret': app_secret,
            'feishu_host': feishu_host,
        }
        keyring.set_password('FeishuApp', app_cfg, json.dumps(feishu_cfg))
        logger.success(f'Feishu App Config for {app_cfg} saved to keyring')
        pass
    else:
        feishu_cfg = json.loads(feishu_cfg)

    sheets = SpreadSheets(
        feishu_cfg['feishu_host'], feishu_cfg['app_id'],
        feishu_cfg['app_secret'], sheet_url)
    sheets.to_excel(output_file)

if __name__ == '__main__':
    get_feishu_table()
